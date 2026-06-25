from pathlib import Path
import sys
import zipfile

from openpyxl import load_workbook


def safe_extract_zip(
    zip_file: Path,
    target_path: Path,
) -> None:
    target_path = target_path.resolve()

    with zipfile.ZipFile(zip_file, "r") as zip_ref:
        for member in zip_ref.infolist():
            member_path = (target_path / member.filename).resolve()

            try:
                member_path.relative_to(target_path)
            except ValueError:
                print(f"Unsafe zip member blocked: {member.filename}")
                sys.exit(1)

        zip_ref.extractall(target_path)


def read_sheet_columns(sheet) -> list[tuple[str, str]]:
    memory_table = []

    for row in sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        column_name = row[0]
        column_type = row[1]

        if column_name is None or column_type is None:
            continue

        column_name = str(column_name).strip()
        column_type = str(column_type).strip().upper()

        if column_name:
            memory_table.append((column_name, column_type))

    return memory_table


def read_credentials_tables(xl_file: Path):
    workbook = load_workbook(
        filename=xl_file,
        read_only=True,
        data_only=True,
    )

    try:
        required_sheet_names = (
            "manufacturer_credentials",
            "product_credentials",
            "document_credentials",
        )
        missing_sheet_names = [
            sheet_name
            for sheet_name in required_sheet_names
            if sheet_name not in workbook.sheetnames
        ]

        if missing_sheet_names:
            missing_text = ", ".join(missing_sheet_names)
            print(f"Required worksheet(s) not found in {xl_file}: {missing_text}")
            sys.exit(1)

        m_table_name = "manufacturer_credentials"
        p_table_name = "product_credentials"
        d_table_name = "document_credentials"
        m_t = read_sheet_columns(workbook[m_table_name])
        p_t = read_sheet_columns(workbook[p_table_name])
        d_t = read_sheet_columns(workbook[d_table_name])
    finally:
        workbook.close()

    return m_table_name, p_table_name, d_table_name, m_t, p_t, d_t


def normalize_folder_name_to_table_base(
    folder_name: str,
    table_suffix: str,
) -> str:
    normalized_name = folder_name.lower().replace(" ", "_")
    return f"{normalized_name}_{table_suffix}"


def get_application_document_table_names(
    applications_path: Path,
    table_suffix: str,
) -> tuple[int, list[str], list[Path]]:
    error_level = 0
    table_names = []
    application_folders = []

    if applications_path.is_dir():
        error_level = 0
    else:
        error_level = 1

    if error_level == 0:
        application_folders = [
            item
            for item in applications_path.iterdir()
            if item.is_dir()
        ]

        if application_folders:
            error_level = 2

    if error_level == 2:
        for folder_path in application_folders:
            table_name = normalize_folder_name_to_table_base(
                folder_name=folder_path.name,
                table_suffix=table_suffix,
            )
            table_names.append(table_name)

    return error_level, table_names, application_folders
